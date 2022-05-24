import openpyxl

class adminLogin:
    i = 5

    #function for showing admin menu
    def showMenu(self):
        while self.i != 4:
            print("1.Add new movie info\n2.Edit movie info\n3.Delete Movie\n4.logout")
            self.i = int(input("Enter choice : "))
            if self.i == 1:
                newMovie = addmovie()
                newMovie.addData()
            elif self.i == 2:
                edit = editMovie()
                edit.editData()
            elif self.i == 3:
                dele = deleteMovie()
                dele.deldata()
            elif self.i == 4:
                break
            else:
                print("Wrong Choice")



# class to add new movie to excel

class addmovie:
    path = r"C:\Users\prajchaudhary\Downloads\movies.xlsx"
    def addData(self):
        wb = openpyxl.load_workbook(self.path)          # loading excel sheet
        sheet = wb.active
        col = sheet.max_column             #get max column
        ro = sheet.max_row                 #get max row
        for j in range(1,col+1):
            #print(f"{sheet.cell(row=1,column=j).value} : ")
            # taking movie info as input for excel sheet
            sheet.cell(row=ro+1, column=j).value = input(f"Enter {sheet.cell(row=1,column=j).value} : ")
        # saving after updation
        wb.save(self.path)

#class to edit movie i.e already present in excel
class editMovie:
    path = r"C:\Users\prajchaudhary\Downloads\movies.xlsx"
    def editData(self):
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.active
        col = sheet.max_column
        ro = sheet.max_row
        print("Select movie to edit")

        # loop for printing movie list
        for j in range(2,ro+1):
            print(f"{j-1}.{sheet.cell(row=j, column=1).value}")
        value = int(input("Enter movie number to be edited: "))

        # loop for updating edited value
        for j in range(1,col+1):
            sheet.cell(row=value+1, column=j).value = input(f"Enter {sheet.cell(row=1,column=j).value} : ")
        wb.save(self.path)

# class for deleting movie
class deleteMovie:
    path = r"C:\Users\prajchaudhary\Downloads\movies.xlsx"
    def deldata(self):
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.active
        col = sheet.max_column
        ro = sheet.max_row
        print("Select movie to delete")
        for j in range(2, ro+1):
            print(f"{j-1}.{sheet.cell(row=j, column=1).value}")
        value = int(input("Enter movie number to be deleted : "))
        sheet.delete_rows(value+1)
        print("Movie deleted successfully")
        wb.save(self.path)