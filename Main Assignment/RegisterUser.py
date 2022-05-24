import openpyxl

# class for registering new user
class registerUser:
    path = r"C:\Users\prajchaudhary\Downloads\idpass.xlsx"
    def registration(self):
        print("*****Create New Account*****")
        wb_obj = openpyxl.load_workbook(self.path)
        sheet_obj = wb_obj.active
        ro = sheet_obj.max_row
        col = sheet_obj.max_column

        #for loop for taking input and write it to excel
        for i in range(1,col+1):
            sheet_obj.cell(row=ro+1, column=i).value = input(f"{sheet_obj.cell(row=1, column=i).value} : ")
        wb_obj.save(self.path)
        print("Registration Successfull")
        print()
