import openpyxl
from admin import adminLogin
from User import userLogin
class loginuser:
    userName = ""
    Password = ""

    # function for taking username and password
    def enter(self):
        self.userName = input("Enter Username : ")
        self.password = input("Enter Password : ")

    # function for validating admin or user according to username and password
    def checkUser(self):
        path = r"C:\Users\prajchaudhary\Downloads\idpass.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active

        for i in range(2,sheet_obj.max_row+1):
            if sheet_obj.cell(row=i, column=1).value == self.userName and sheet_obj.cell(row=i, column=2).value == self.password:
                if i == 2:
                    print("1.Admin Login")
                    print("******Welcome Admin******* ")
                    admin_login = adminLogin()
                    admin_login.showMenu()
                    break
                else:
                    print("2.User Login ")
                    print("***** Welcome to Book My Show *****")
                    print()
                    print(f"***** Welcome {sheet_obj.cell(row=i, column=1).value} *****")
                    user1 = userLogin()
                    user1.showmovielist()
                    break