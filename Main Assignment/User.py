import openpyxl

class userLogin:
    path = r"C:\Users\prajchaudhary\Downloads\movies.xlsx"

    #function for showing title of every movie
    def showmovielist(self):
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.active
        col = sheet.max_column
        ro = sheet.max_row
        value = 100
        while value != 0:
            for j in range(2, ro + 1):
                print(f"{j - 1}.{sheet.cell(row=j, column=1).value}")
            print("0 -> logout")
            value = int(input("Enter option : "))
            if value == 0:
                break
            for i in range(1, col + 1):
                print(f"{sheet.cell(row=1, column=i).value} : {sheet.cell(row=value + 1, column=i).value}")

            print()
            print("1.Book Ticket\n2.Cancel Ticket\n3.Give user rating\n")
            choose = int(input("Select Option : "))
            if choose == 1:
                book = bookTicket()
                book.ticketbook(value + 1)
            if choose == 2:
                cancel = cancelTicket()
                cancel.ticketcancel(value + 1)
            if choose == 3:
                rating = userRating()
                rating.ratings(value + 1)



# class for book movie ticket

class bookTicket:
    remSeats = 0
    path = r"C:\Users\prajchaudhary\Downloads\movies.xlsx"
    def ticketbook(self,value):
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.active
        self.remSeats = int(sheet.cell(row=value, column=11).value)
        print(f"Remaining Seats : {self.remSeats}")
        seatsTOBook = int(input("Enter number of seats : "))
        if(seatsTOBook > self.remSeats):
            print("Limit Exceed")
        else:
            self.remSeats = self.remSeats - seatsTOBook
            sheet.cell(row=value,column=11).value = self.remSeats
            print("Thanks for Booking")
        wb.save(self.path)

# class for cancel movie ticket
class cancelTicket:
    path = r"C:\Users\prajchaudhary\Downloads\movies.xlsx"
    ticketstobecancel = 0
    def ticketcancel(self,value):
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.active
        self.ticketstobecancel = int(input("No of seats to be cancel : "))
        rem = int(sheet.cell(row=value, column=11).value)
        if(rem < self.ticketstobecancel):
            print("Limit exceed")
        else:
            sheet.cell(row=value, column=11).value = rem + self.ticketstobecancel
            print("Tickets Successfully Canceled")
        wb.save(self.path)

# class for user rating
class userRating:
    def ratings(self,value):
        rat = int(input("Enter rating for following movie : "))
